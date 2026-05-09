"""
Bootstrap Import Wizard — pure import service functions.

All functions in this module are pure (no side effects, no network calls).
The router (import_wizard.py) is responsible for I/O and session management.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import re
import uuid
from dataclasses import dataclass, field
from typing import Any

import openpyxl  # type: ignore

logger = logging.getLogger(__name__)

_SECTION_RE = re.compile(r"^##(\w+)", re.MULTILINE)
_DATE_YYYYMMDD_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_DATE_MMDDYYYY_RE = re.compile(r"^(\d{2})/(\d{2})/(\d{4})$")


class ImportParseError(Exception):
    """Raised for CSV structure violations (missing sections, no markers)."""


KNOWN_ENUM_MAPS: dict[str, dict[str, str]] = {
    "Shot_Eligibility": {
        "OK with Milk": "Passable",
        "Passable Espresso": "Passable",
        "Good": "Good Espresso",
        "Reject": "Reject",
        "Good Espresso": "Good Espresso",
        "God Shot": "God Shot",
    },
    "Taste_Summary": {
        "Weak and Sour": "Weak & Sour",
        "Balanced and Neutral": "Sweet & Balanced",
        "Too acidic": "Acidic & Bright",
        "Too bitter": "Harsh & Bitter",
    },
    "Roast_Level": {
        "Medium-Dark": "Medium / Dark",
        "Light/Medium": "Light / Medium",
    },
}

CANONICAL_ENUM_VALUES: dict[str, frozenset[str]] = {
    "Shot_Eligibility": frozenset({"Reject", "Passable", "Good Espresso", "God Shot"}),
    "Taste_Summary": frozenset(
        {
            "Weak & Sour",
            "Acidic & Bright",
            "Sweet & Balanced",
            "Complex & Syrupy",
            "Harsh & Bitter",
            "Strong & Muddy",
            "Salty / Channeled",
        }
    ),
    "Roast_Level": frozenset({"Light", "Light / Medium", "Medium", "Medium / Dark", "Dark"}),
}

CANONICAL_COLUMNS: dict[str, list[str]] = {
    "Brew_Log": [
        "Shot_ID",
        "Date",
        "Bag_ID",
        "Machine_ID",
        "Grinder_ID",
        "Basket_ID",
        "Dose_In_g",
        "Yield_Out_g",
        "Time_Sec",
        "Grind_Setting",
        "Shot_Eligibility",
        "Taste_Summary",
        "User_Notes",
        "AI_Feedback",
        "Storage_Method",
    ],
    "Catalog": [
        "Catalog_ID",
        "Roaster",
        "Bean_Name",
        "Roast_Level",
        "Product_URL",
        "Local_Image_Path",
    ],
    "Hardware": ["Hardware_ID", "Category", "Name"],
    "Inventory": [
        "Bag_ID",
        "Beans",
        "RoastDate",
        "RoastLevel",
        "Display_Name",
        "Catalog_ID",
        "Status",
        "Storage_Method",
    ],
    "Maintenance": ["Maintenance_ID", "Hardware_ID", "Date", "Action_Type", "Notes"],
}


def parse_legacy_csv(content: str) -> dict[str, list[dict[str, Any]]]:
    """Split multi-section CSV on ##SheetName markers.

    Raises ImportParseError if no section markers or ##Brew_Log is absent.
    """
    content = content.replace("\r\n", "\n").replace("\r", "\n")
    markers = list(_SECTION_RE.finditer(content))
    if not markers:
        raise ImportParseError("No sections found in CSV")
    sections: dict[str, list[dict[str, Any]]] = {}
    for i, match in enumerate(markers):
        name = match.group(1)
        start = match.end() + 1
        end = markers[i + 1].start() if i + 1 < len(markers) else len(content)
        block = content[start:end].strip()
        if not block:
            sections[name] = []
            continue
        reader = csv.DictReader(io.StringIO(block))
        sections[name] = [dict(row) for row in reader]
    if "Brew_Log" not in sections:
        raise ImportParseError("Required section Brew_Log is missing")
    return sections


def parse_xlsx_to_sections(raw: bytes) -> dict[str, list[dict[str, Any]]]:
    """Parse an XLSX file into sections matching parse_legacy_csv's output format.

    Each worksheet tab name becomes a section key. The first non-empty row of
    each sheet is treated as headers; subsequent rows become row dicts.

    Raises ImportParseError if the workbook is invalid, has no sheets, or lacks Brew_Log.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as exc:
        raise ImportParseError(f"Could not open XLSX file: {exc}") from exc
    try:
        sheet_names = wb.sheetnames
        if not sheet_names:
            raise ImportParseError("XLSX file contains no sheets")

        sections: dict[str, list[dict[str, Any]]] = {}
        for name in sheet_names:
            ws = wb[name]
            all_rows = list(ws.iter_rows(values_only=True))

            # Find first non-empty row to use as headers
            header_idx = None
            for i, row in enumerate(all_rows):
                if any(v is not None and str(v).strip() != "" for v in row):
                    header_idx = i
                    break

            if header_idx is None:
                sections[name] = []
                continue

            raw_headers = [str(v).strip() if v is not None else "" for v in all_rows[header_idx]]
            # Trim trailing empty/unnamed headers (e.g. AppSheet exports wide sheets)
            last_named = max((i for i, h in enumerate(raw_headers) if h), default=-1)
            headers = raw_headers[: last_named + 1] if last_named >= 0 else raw_headers

            def _cell_str(v: object) -> str:
                """Convert a cell value to string; collapse whole-number floats (17.0→'17')."""
                if v is None:
                    return ""
                if isinstance(v, float) and v == int(v):
                    return str(int(v))
                return str(v)

            rows: list[dict[str, Any]] = []
            for row in all_rows[header_idx + 1 :]:
                # Skip completely empty rows
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                row_dict = {
                    headers[i]: _cell_str(v)
                    for i, v in enumerate(row)
                    if i < len(headers) and headers[i]  # skip empty-keyed columns
                }
                rows.append(row_dict)

            sections[name] = rows

        if "Brew_Log" not in sections:
            raise ImportParseError("Required sheet Brew_Log is missing")

        # Drop AppSheet pre-populated "template" rows from Brew_Log: these are rows
        # where Machine_ID/Grinder_ID are pre-filled by AppSheet for future use but
        # Date and Bag_ID are empty — they are not actual brew log entries.
        brew = sections["Brew_Log"]
        sections["Brew_Log"] = [
            r for r in brew if r.get("Date", "").strip() and r.get("Bag_ID", "").strip()
        ]

        return sections
    finally:
        wb.close()


def build_mapping_prompt(
    sheet_name: str,
    legacy_columns: list[str],
    canonical_columns: list[str],
    sample_rows: list[dict[str, Any]],
) -> str:
    """Build the LLM header-mapping prompt (<=1024 bytes encoded).

    NOT called for Grinder_Calibration.
    """
    rows = sample_rows[:3]
    for attempt in range(len(rows) + 1):
        truncated = [
            {k: str(v)[:80] for k, v in row.items()} for row in rows[: max(1, len(rows) - attempt)]
        ]
        prompt = (
            "You are mapping legacy database columns to a canonical schema.\n\n"
            f"Sheet: {sheet_name}\n"
            f"Legacy columns: {json.dumps(legacy_columns)}\n"
            f"Canonical columns: {json.dumps(canonical_columns)}\n"
            "Sample rows (first 3, values truncated to 80 chars):\n"
            f"{json.dumps(truncated, indent=2)}\n\n"
            "Return ONLY a JSON object in this exact format — no prose, no markdown fences:\n"
            '{"mapping": {"<legacy_col>": "<canonical_col_or_null>", ...}}\n\n'
            "Rules:\n"
            "- Map every legacy column to the most likely canonical column name.\n"
            '- Use JSON null (not the string "null") when no canonical column matches.\n'
            "- Do not invent canonical column names; use only those in the Canonical columns list."
        )
        if len(prompt.encode()) <= 1024:
            return prompt
    return prompt


def parse_mapping_response(llm_response: str) -> dict[str, str | None]:
    """Extract {legacy_col: canonical_col_or_None} from LLM JSON.

    Returns {} on any error.
    """
    try:
        data = json.loads(llm_response)
        return {k: v for k, v in data["mapping"].items()}
    except (json.JSONDecodeError, KeyError, TypeError) as exc:
        logger.warning("Failed to parse LLM mapping response: %s", exc)
        return {}


def build_batch_mapping_prompt(
    sections: dict[str, list[dict[str, Any]]],
) -> str:
    """Build a single LLM prompt mapping all sections at once (one API call).

    Replaces N separate build_mapping_prompt calls with one combined prompt.
    Skips Grinder_Calibration and empty sections.
    """
    lines = [
        "You are mapping legacy spreadsheet columns to canonical schema columns.",
        "Map each section's legacy columns to the canonical columns provided.",
        "",
        "Return ONLY a JSON object — no prose, no markdown fences — in this exact format:",
        '{"sections": {"<section>": {"mapping": {"<legacy_col>": "<canonical_col_or_null>"}}, ...}}',
        "",
        "Rules:",
        "- Map every legacy column to the most likely canonical column name.",
        '- Use JSON null (not the string "null") when no canonical column matches.',
        "- Do not invent canonical column names; use only those listed per section.",
        "",
    ]

    for section_name, rows in sections.items():
        if section_name == "Grinder_Calibration" or not rows:
            continue
        canonical_cols = CANONICAL_COLUMNS.get(section_name, [])
        legacy_cols = list(rows[0].keys())
        sample = [{k: str(v)[:60] for k, v in row.items()} for row in rows[:2]]
        lines.append(f"## {section_name}")
        lines.append(f"  Canonical columns: {json.dumps(canonical_cols)}")
        lines.append(f"  Legacy columns: {json.dumps(legacy_cols)}")
        lines.append(f"  Sample rows: {json.dumps(sample)}")
        lines.append("")

    return "\n".join(lines)


def parse_batch_mapping_response(
    llm_response: str,
    sections: dict[str, list[dict[str, Any]]],
) -> dict[str, dict[str, str | None]]:
    """Extract per-section mappings from a batched LLM response.

    Falls back to identity mapping for any section that can't be parsed.
    """
    results: dict[str, dict[str, str | None]] = {}
    try:
        data = json.loads(llm_response)
        section_data = data.get("sections", {})
    except (json.JSONDecodeError, TypeError) as exc:
        logger.warning("Failed to parse batch LLM mapping response: %s", exc)
        section_data = {}

    for section_name, rows in sections.items():
        if section_name == "Grinder_Calibration" or not rows:
            continue
        legacy_cols = list(rows[0].keys())
        raw = section_data.get(section_name, {})
        mapping = raw.get("mapping", {}) if isinstance(raw, dict) else {}
        if not mapping:
            logger.warning("No mapping returned for section %s — using identity", section_name)
            results[section_name] = {col: col for col in legacy_cols}
        else:
            results[section_name] = {k: v for k, v in mapping.items()}
    return results


def find_enum_divergences(
    rows: list[dict[str, Any]],
    column_mapping: dict[str, str | None],
) -> dict[str, list[str]]:
    """Find legacy enum values that cannot be auto-resolved.

    Only inspects columns whose canonical target is a key in KNOWN_ENUM_MAPS.
    """
    enum_cols: dict[str, str] = {
        legacy: canonical
        for legacy, canonical in column_mapping.items()
        if canonical in KNOWN_ENUM_MAPS
    }
    divergences: dict[str, set[str]] = {}
    for row in rows:
        for legacy_col, canonical_col in enum_cols.items():
            val = row.get(legacy_col, "")
            if not val:
                continue
            known_map = KNOWN_ENUM_MAPS[canonical_col]
            canonical_vals = CANONICAL_ENUM_VALUES.get(canonical_col, frozenset())
            if val not in known_map and val not in canonical_vals:
                divergences.setdefault(canonical_col, set()).add(val)
    return {k: sorted(v) for k, v in divergences.items() if v}


def normalize_brew_log_row(
    row: dict[str, Any], column_mapping: dict[str, Any], confirmed_enum_maps: dict[str, Any]
) -> dict[str, Any]:
    """Rename legacy keys -> canonical; discard None-mapped; normalize date + enums."""
    out: dict[str, Any] = {}
    for legacy_key, value in row.items():
        canonical_key = column_mapping.get(legacy_key)
        if canonical_key is None:
            continue
        out[canonical_key] = value
    if "Date" in out:
        date_val = out["Date"]
        if _DATE_YYYYMMDD_RE.match(date_val):
            pass
        else:
            m = _DATE_MMDDYYYY_RE.match(date_val)
            if m:
                mm, dd, yyyy = m.group(1), m.group(2), m.group(3)
                out["Date"] = f"{yyyy}-{mm}-{dd}"
            else:
                logger.warning(
                    "Unrecognized date format %r in Brew_Log row; writing as-is", date_val
                )
    for enum_col in ("Shot_Eligibility", "Taste_Summary"):
        if enum_col not in out:
            continue
        val = out[enum_col]
        resolved = KNOWN_ENUM_MAPS.get(enum_col, {}).get(val)
        if resolved is None:
            resolved = confirmed_enum_maps.get(enum_col, {}).get(val)
        if resolved is not None:
            out[enum_col] = resolved
    # Auto-generate Shot_ID when absent from legacy schema
    if not out.get("Shot_ID"):
        out["Shot_ID"] = f"SHOT-{uuid.uuid4().hex[:8].upper()}"
    return out


def normalize_catalog_row(
    row: dict[str, Any], column_mapping: dict[str, Any], confirmed_enum_maps: dict[str, Any]
) -> dict[str, Any]:
    """Rename legacy keys -> canonical; normalize Roast_Level enum."""
    out = {
        canonical: row[legacy]
        for legacy, canonical in column_mapping.items()
        if canonical is not None and legacy in row
    }
    if "Roast_Level" in out:
        val = out["Roast_Level"]
        resolved = KNOWN_ENUM_MAPS.get("Roast_Level", {}).get(val)
        if resolved is None:
            resolved = confirmed_enum_maps.get("Roast_Level", {}).get(val)
        if resolved is not None:
            out["Roast_Level"] = resolved
    return out


def normalize_inventory_row(
    row: dict[str, Any], column_mapping: dict[str, Any], confirmed_enum_maps: dict[str, Any]
) -> dict[str, Any]:
    """Rename legacy keys -> canonical; normalize RoastLevel enum.

    Note: Legacy column is 'RoastLevel'; canonical column is also 'RoastLevel'.
    """
    out = {
        canonical: row[legacy]
        for legacy, canonical in column_mapping.items()
        if canonical is not None and legacy in row
    }
    if "RoastLevel" in out:
        val = out["RoastLevel"]
        resolved = KNOWN_ENUM_MAPS.get("Roast_Level", {}).get(val)
        if resolved is None:
            resolved = confirmed_enum_maps.get("Roast_Level", {}).get(val)
        if resolved is not None:
            out["RoastLevel"] = resolved
    return out


def normalize_hardware_row(
    row: dict[str, Any], column_mapping: dict[str, Any], confirmed_enum_maps: dict[str, Any]
) -> dict[str, Any]:  # noqa: ARG001
    """Rename legacy keys -> canonical (no enum normalization for Hardware)."""
    return {
        canonical: row[legacy]
        for legacy, canonical in column_mapping.items()
        if canonical is not None and legacy in row
    }


def migrate_grinder_calibration_row(row: dict[str, Any], sequence: int) -> dict[str, Any]:
    """Convert a legacy Grinder_Calibration row to a Maintenance row."""
    return {
        "Maintenance_ID": f"MNT{sequence:03d}",
        "Hardware_ID": row["Grinder_ID"],
        "Date": row["Date"],
        "Action_Type": "Re-zero",
        "Notes": row.get("Notes", ""),
    }


@dataclass
class ImportState:
    """Holds the full wizard state serialised into the session cookie."""

    sections: dict[str, list[dict[str, Any]]]
    column_mappings: dict[str, dict[str, Any]]
    enum_divergences: dict[str, list[str]]
    confirmed_enum_maps: dict[str, dict[str, Any]]
    dry_run_preview: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
